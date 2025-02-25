from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import pandas as pd
import numpy as np
from thefuzz import fuzz
import nltk
import string
from nltk.corpus import stopwords
from nltk.stem import WordNetLemmatizer
from sentence_transformers import SentenceTransformer, util
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
import os
import time
import shutil

# Additional imports for formatting
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

app = Flask(__name__)
CORS(app)

# Download necessary NLTK data files (only needed once)
nltk.download('stopwords')
nltk.download('wordnet')
nltk.download('omw-1.4')

# Initialize Sentence Transformer with a better model
model = SentenceTransformer('all-mpnet-base-v2')

# Initialize Lemmatizer and Stop Words
lemmatizer = WordNetLemmatizer()
stop_words = set(stopwords.words('english'))

UPLOAD_FOLDER = 'uploads'
OUTPUT_FOLDER = 'output'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

def preprocess_text(text):
    if pd.isnull(text):
        return ''
    text = str(text).lower()
    text = text.translate(str.maketrans('', '', string.punctuation))
    words = text.split()
    words = [lemmatizer.lemmatize(word) for word in words]
    return ' '.join(words)

@app.route('/process', methods=['POST'])
def process_files():
    start_time = time.time()
    print("Received request to process files")

    # Clear output folder before saving new results
    shutil.rmtree(OUTPUT_FOLDER)
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

    # Get uploaded files and Top-K value
    file1 = request.files.get('frame1')
    file2 = request.files.get('frame2')
    top_k = int(request.form.get('top_k', 6))  # Default to 6 if not provided
    print(f"Top-K value set to: {top_k}")

    if not file1 or not file2:
        print("Error: Both files are required")
        return jsonify({"error": "Both files are required"}), 400

    # Save uploaded files
    path1 = os.path.join(UPLOAD_FOLDER, file1.filename)
    path2 = os.path.join(UPLOAD_FOLDER, file2.filename)
    file1.save(path1)
    file2.save(path2)
    print(f"Files saved: {path1}, {path2}")

    # Load the Excel Files
    df1 = pd.read_excel(path1)
    df2 = pd.read_excel(path2)

    # Forward-fill missing 'Domain' and 'Sub-Domain'
    df1['Domain'] = df1['Domain'].fillna(method='ffill')
    df2['Domain'] = df2['Domain'].fillna(method='ffill')
    df1['Sub-Domain'] = df1['Sub-Domain'].fillna(method='ffill')
    df2['Sub-Domain'] = df2['Sub-Domain'].fillna(method='ffill')

    # Combine and Preprocess Text
    df1['Combined_Text'] = df1['Domain'].astype(str) + ' ' + df1['Sub-Domain'].astype(str) + ' ' + df1['Control'].astype(str)
    df2['Combined_Text'] = df2['Domain'].astype(str) + ' ' + df2['Sub-Domain'].astype(str) + ' ' + df2['Control'].astype(str)

    df1['Processed_Control'] = df1['Combined_Text'].apply(preprocess_text)
    df2['Processed_Control'] = df2['Combined_Text'].apply(preprocess_text)

    # Match Domains Using Semantic Similarity
    domains1 = df1['Domain'].dropna().unique().tolist()
    domains2 = df2['Domain'].dropna().unique().tolist()

    # Manual domain mappings
    manual_domain_mapping = {
        'Security Incident Management': 'Resilience',
        'Application & Software': 'Secure Software Development Lifecycle',
    }

    # Domain descriptions
    domain_descriptions = {
        'Security Incident Management': 'Management of security incidents and response',
        'Resilience': 'Organizational resilience and disaster recovery',
        'Application & Software': 'Development and management of applications and software',
        'Secure Software Development Lifecycle': 'Practices for secure software development',
    }

    domain_texts1 = [preprocess_text(domain_descriptions.get(d, d)) for d in domains1]
    domain_texts2 = [preprocess_text(domain_descriptions.get(d, d)) for d in domains2]

    domain_embeddings1 = model.encode(domain_texts1, convert_to_tensor=True)
    domain_embeddings2 = model.encode(domain_texts2, convert_to_tensor=True)

    domain_similarity_matrix = util.cos_sim(domain_embeddings1, domain_embeddings2).cpu().numpy()

    fuzzy_scores = []
    for d1 in domains1:
        scores = [fuzz.token_set_ratio(d1, d2) for d2 in domains2]
        fuzzy_scores.append(scores)
    fuzzy_similarity_matrix = np.array(fuzzy_scores) / 100

    combined_similarity_matrix = (domain_similarity_matrix * 0.6 + fuzzy_similarity_matrix * 0.4)
    domain_similarity_threshold = 0.44

    domain_mapping = {}
    for idx1, d1 in enumerate(domains1):
        if d1 in manual_domain_mapping:
            domain_mapping[d1] = manual_domain_mapping[d1]
            print(f"Manual mapping: {d1} --> {domain_mapping[d1]}")
        else:
            similarities = combined_similarity_matrix[idx1]
            best_idx = similarities.argmax()
            best_score = similarities[best_idx]
            if best_score >= domain_similarity_threshold:
                d2 = domains2[best_idx]
                domain_mapping[d1] = d2
                print(f"Matched {d1} --> {d2} (Similarity: {best_score:.2f})")
            else:
                print(f"No match for Domain: {d1} (Best score: {best_score:.2f})")

    print("\nFinal Matched Domains:")
    for d1, d2 in domain_mapping.items():
        print(f"{d1} --> {d2}")

    # Match Controls
    results = []
    tfidf_weight = 0.4
    embedding_weight = 0.6
    control_threshold = 0.35

    for domain_f1, domain_f2 in domain_mapping.items():
        controls_f1 = df1[df1['Domain'] == domain_f1].reset_index(drop=True)
        controls_f2 = df2[df2['Domain'] == domain_f2].reset_index(drop=True)

        if controls_f1.empty or controls_f2.empty:
            print(f"No controls found for domain '{domain_f1}' or '{domain_f2}'. Fallback to all domains.")
            controls_f1 = df1.copy()
            controls_f2 = df2.copy()
            domain_f1 = 'All Domains'

        texts_f1 = controls_f1['Processed_Control'].tolist()
        texts_f2 = controls_f2['Processed_Control'].tolist()

        embeddings_f1 = model.encode(texts_f1, convert_to_tensor=True)
        embeddings_f2 = model.encode(texts_f2, convert_to_tensor=True)
        embedding_similarity = util.cos_sim(embeddings_f1, embeddings_f2).cpu().numpy()

        vectorizer = TfidfVectorizer().fit(texts_f1 + texts_f2)
        tfidf_f1 = vectorizer.transform(texts_f1)
        tfidf_f2 = vectorizer.transform(texts_f2)
        tfidf_similarity = cosine_similarity(tfidf_f1, tfidf_f2)

        combined_similarity = (embedding_similarity * embedding_weight +
                               tfidf_similarity * tfidf_weight) / (embedding_weight + tfidf_weight)

        for idx_f1, row_f1 in controls_f1.iterrows():
            similarities = combined_similarity[idx_f1]
            top_indices = similarities.argsort()[-top_k:][::-1]
            top_scores = similarities[top_indices]
            matching_indices = [i for i, score in zip(top_indices, top_scores) if score >= control_threshold]
            matching_controls = [controls_f2.iloc[i]['Control'] for i in matching_indices]
            matching_scores = [similarities[i] for i in matching_indices]

            if not matching_controls:
                matching_controls = [None]
                matching_scores = [None]

            results.append({
                'Domain': domain_f1,
                'Sub-Domain': row_f1['Sub-Domain'],
                'Control': row_f1['Control'],
                'Controls from F2 that match': matching_controls,
                'Similarity Scores': matching_scores
            })

    final_df = pd.DataFrame(results)
    final_df = final_df.explode(['Controls from F2 that match', 'Similarity Scores'])
    final_df = final_df.reset_index(drop=True)
    final_df['Similarity Scores'] = final_df['Similarity Scores'].apply(lambda x: f"{x:.2f}" if pd.notnull(x) else '')

    final_df['Sr No.'] = final_df.index + 1
    final_df = final_df.rename(columns={
        'Domain': 'User Org Domain',
        'Sub-Domain': 'User Org Sub-Domain',
        'Control': 'User Org Control Statement',
        'Controls from F2 that match': 'Matched Control from Service Org Framework',
        'Similarity Scores': 'Similarity Score'
    })

    final_df = final_df[['Sr No.', 'User Org Domain', 'User Org Sub-Domain', 'User Org Control Statement',
                         'Matched Control from Service Org Framework', 'Similarity Score']]

    output_path = os.path.join(OUTPUT_FOLDER, 'framework1_with_results.xlsx')
    final_df.to_excel(output_path, index=False)

    # Formatting with openpyxl
    wb = load_workbook(output_path)
    ws = wb.active

    # Bold headers
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Wrap text for all cells
    wrap_alignment = Alignment(wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = wrap_alignment

    # Set column widths based only on header length
    # Add some padding to ensure readability
    for i, col_name in enumerate(final_df.columns, start=1):
        header_length = len(col_name)
        ws.column_dimensions[get_column_letter(i)].width = header_length + 5

    wb.save(output_path)

    print(f"Results saved to {output_path}")
    total_time = time.time() - start_time
    print(f"Total processing time: {total_time:.2f} seconds")

    return jsonify({
        "data": final_df.to_dict(orient='records'),
        "processing_time": f"{total_time:.2f} seconds"
    }), 200

# Updated download route so that the file is served directly from the base URL
@app.route('/framework1_with_results.xlsx', methods=['GET'])
def download_file():
    output_path = os.path.join(OUTPUT_FOLDER, 'framework1_with_results.xlsx')
    if not os.path.exists(output_path):
        print("Error: File not found for download")
        return jsonify({"error": "File not found"}), 404
    print(f"File downloaded: {output_path}")
    response = send_file(output_path, as_attachment=True)
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5001, threaded=True, debug=False)
